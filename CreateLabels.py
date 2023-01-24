import openpyxl
import os

from ctypes import windll
from tkinter import filedialog
import blabel

def get_data(path):
    # Open the Excel file and get the data from the first sheet
    wb = openpyxl.load_workbook(path)

    # Get the first sheet
    sheet = wb.active

    # Iterate over the rows in the sheet and yield the data
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        yield Item(row[0], row[1], row[2], row[15])

class Item:
    def __init__(self, inventory, item_id, description, transfer):
        self.item_id = item_id
        self.description = description
        self.transfer = transfer
        self.inventory = inventory
    
    def __str__(self) -> str:
        return f'{self.item_id} - {self.description}'

# Write a main function to test the code
def main():
     # Open file explorer and get the path to the Excel file
    path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    print(path)

    # Set the DPI awareness to Per Monitor v2
    windll.shcore.SetProcessDpiAwareness(1)
    
    # Create a LabelWriter object
    label_writer = blabel.LabelWriter("label_template.html", default_stylesheets=("style.css",))

    # Get the data from the Excel file
    label_count = 1
    records = []
    record = {}
    for item in get_data(path):
        print(item)
        # If the label count is 30, add the record to the list of records
        if label_count > 30:
            records.append(record)
            # Reset the label count and the record
            label_count = 1
            record = {}
        # Add the data to the record
        record['item_id_' + str(label_count)] = item.item_id
        record['description_' + str(label_count)] = item.description
        if item.transfer == 'N' and item.inventory != 'MORDEP':
            record['symbol_' + str(label_count)] = '\u25C6'
            record['color_' + str(label_count)] = 'red'
        elif item.transfer == 'Y' and item.inventory == 'MORDEP':
            record['symbol_' + str(label_count)] = '\u25C7'
            record['color_' + str(label_count)] = 'blue'
        elif item.transfer == 'N' and item.inventory == 'MORDEP':
            record['symbol_' + str(label_count)] = '\u25C8'
            record['color_' + str(label_count)] = 'purple'
        else:
            record['symbol_' + str(label_count)] = ''
            record['color_' + str(label_count)] = ''

        label_count += 1
    if label_count > 0:
        records.append(record)
    # Write the labels to a PDF file
    file_name = os.path.basename(os.path.splitext(path)[0])
    file_name = f'{file_name}.pdf'
    label_writer.write_labels(records, target=file_name)

if __name__ == '__main__':
    main()
